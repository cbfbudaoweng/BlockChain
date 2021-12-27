# -*- coding: UTF-8 -*-
# @Author  : cbf
# @time    : 2021/6/12 10:28
# @File    : getExcelData.py
# @Software: PyCharm

from openpyxl import load_workbook
from common.baseapi import BaseApi


class OpenpyXl(BaseApi):
    # 读excel
    # def read_excel_data(self, filename, sheetName, endRow, startRow, requestCol=7, expectedCol=9):
    def read_excel_data(self, filename, sheetName, endRow, startRow, requestCol=12, expectedCol=13):
        """
        :param filename: 参数传入excel文件名
        :param sheetName: 参数传入excel文件需要操作的sheet名
        :param endRow: 参数传入excel文件需要操作的用例结束行数
        :param statRow: 参数传入excel文件需要操作的用例开始行数
        :param requestCol: 关键字参数传入excel文件需要操作的用例请求参数列，由于有用例模块，所以请求参数列就是第7行，比较固定，暂不需要传入
        :param expectedCol: 关键字参数传入excel文件需要操作的用例预期返回数据列，由于有用例模块，所以预期返回数据列就是第9行，比较固定，暂不需要传入
        :return:datalist列表--[(用例编号)，(请求参数),(预期返回数据)]
        """
        datalist = []  # 存放数据,存放格式--[(用例编号)，(请求参数),(预期返回数据)]
        wb = load_workbook(filename)  # 使用openpyxl读取xlsx文件，创建workbook
        # sheets=wb.sheetnames  # 获取xlsx文件中所有的sheet
        # pprint(sheets)  # 打印所有的sheets
        # print(wb[sheet])  # 打印指定的sheet
        apisheet = wb[sheetName]  # 把指定的sheet存到apisheet
        # myfile = self.create_file('../res ult/getdata.txt')  # 调用file_operations方法，生成文件用于存放输出的内容
        # 读取excel中的请求参数列和预期返回数据列，并存放到datalist中
        for i in range(startRow, endRow + 1):  # 左含右不含[2，6) 即取第 2，3，4，5行的用例编号、请求参数、预期返回数据列
            datalist.append((apisheet.cell(i, 1).value, apisheet.cell(i, requestCol).value, apisheet.cell(i, expectedCol).value))  # 将输出信息存到datalist中
        wb.close()  # 关闭文件
        print(datalist)
        return datalist  # [(),(),()] 列表内套元组
        #     case_num = apisheet.cell(i, 1).value  # 获取用例中的用例编号
        #     requests_data = apisheet.cell(i, requestCol).value  # 获取用例中的请求参数
        #     expected_data = apisheet.cell(i, expectedCol).value  # 获取用例中的预期返回数据
        #     print(case_num, requests_data, expected_data, file=myfile)  # 将输出的用例编号、请求参数、预期返回数据信息写入到文件中
        # myfile.close()  # 关闭文件
        # wb.close()  # 关闭文件

    # 写excel
    def write_excel_data(self, filename, sheetName, writeRow, writeCol, writeData):
        """
        :param filename: 参数传入excel文件名
        :param sheetName: 参数传入excel文件需要操作的sheet名
        :param writeRow: 参数传入excel文件需要操作的用例写入行
        :param writeCol: 参数传入excel文件需要操作的用例写入列
        :param writeData: 参数传入excel文件需要写入的内容
        :return:
        """
        wb = load_workbook(filename)  # 使用openpyxl读取xlsx文件，创建workbook
        apisheet = wb[sheetName]  # 把指定的sheet存到apisheet
        # # 把值写入到excel的实际返回数据或测试结果列中，即第10、11列
        # for i in range(startRow, endRow + 1):  # 左含右不含[2，6) 即取第 2，3，4，5行的第writeCol列写入数据
        #     apisheet.cell(i, writeCol).value = writeData
        #     wb.save(filename)  # 保存excel文件
        # wb.save(filename)  # 保存excel文件
        # wb.close()  # 关闭
        apisheet.cell(writeRow, writeCol).value = writeData  # 在writeRow行writeCol列，写入内容writeData
        wb.save(filename)  # 保存excel文件
        wb.close()  # 关闭文件
